import customtkinter as ctk
from PIL import Image, ImageTk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta
import os
from win10toast import ToastNotifier
from openpyxl import Workbook, load_workbook
import shutil
import multiprocessing as mp
import time

# Constants
EXCEL_FILE = "data.xlsx"
USERS_SHEET = "Users"
DATA_FOLDER = "user_data"
IMAGE_SIZE_PREVIEW = (300, 150)
IMAGE_SIZE_CARD = (250, 100)
IMAGE_SIZE_DETAIL = (320, 150)

class UserManager:
    def __init__(self):
        self.users = {}
        self.current_user = None
        os.makedirs(DATA_FOLDER, exist_ok=True)
        self.load_users_from_excel()

    def load_users_from_excel(self):
        if not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = load_workbook(EXCEL_FILE)
            if USERS_SHEET in wb.sheetnames:
                for row in wb[USERS_SHEET].iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        self.users[row[0]] = row[1]
        except:
            pass

    def save_user_to_excel(self, username, password):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = USERS_SHEET
            ws.append(["Username", "Password"])
            wb.save(EXCEL_FILE)

        wb = load_workbook(EXCEL_FILE)
        if USERS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(USERS_SHEET)
            ws.append(["Username", "Password"])
        else:
            ws = wb[USERS_SHEET]

        ws.append([username, password])
        wb.save(EXCEL_FILE)

    def register_user(self, username, password, confirm_password):
        if password != confirm_password:
            raise ValueError("Password tidak cocok")
        if username in self.users:
            raise ValueError("Username sudah dipakai")
        self.users[username] = password
        self.save_user_to_excel(username, password)
        os.makedirs(os.path.join(DATA_FOLDER, username), exist_ok=True)
        return True

    def login_user(self, username, password):
        if username not in self.users:
            raise ValueError("Username tidak ditemukan")
        if self.users[username] != password:
            raise ValueError("Password salah")
        self.current_user = username
        current_user_file = os.path.join(DATA_FOLDER, "current_user.txt")
        with open(current_user_file, 'w') as f:
            f.write(username)
        return True

    def logout_user(self):
        self.current_user = None
        current_user_file = os.path.join(DATA_FOLDER, "current_user.txt")
        if os.path.exists(current_user_file):
            os.remove(current_user_file)

class SavingsManager:
    def __init__(self, user_manager):
        self.user_manager = user_manager
        self.savings = []
        self.image_refs = {}

    def get_user_images_folder(self):
        if not self.user_manager.current_user:
            return None
        folder = os.path.join(DATA_FOLDER, self.user_manager.current_user, "images")
        os.makedirs(folder, exist_ok=True)
        return folder

    def create_user_sheet(self, username):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = USERS_SHEET
            ws.append(["Username", "Password"])
            wb.save(EXCEL_FILE)
        wb = load_workbook(EXCEL_FILE)
        if username not in wb.sheetnames:
            ws = wb.create_sheet(username)
            ws.append(["Nama","Target","Nominal","Rencana","Estimasi Hari",
                       "Estimasi Minggu","Estimasi Bulan","Gambar Path",
                       "Terkumpul","Notifikasi","Notif Aktif","Tanggal Dibuat"])
            wb.save(EXCEL_FILE)

    def load_user_data(self):
        self.savings = []
        if not os.path.exists(EXCEL_FILE) or not self.user_manager.current_user:
            return
        try:
            wb = load_workbook(EXCEL_FILE)
            if self.user_manager.current_user not in wb.sheetnames:
                self.create_user_sheet(self.user_manager.current_user)
                return
            ws = wb[self.user_manager.current_user]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    self.savings.append({
                        "nama": row[0], "target": self.to_int(row[1]), "nominal": self.to_int(row[2]),
                        "rencana": row[3] or "Harian", "estimasi_hari": self.to_int(row[4]),
                        "estimasi_minggu": self.to_int(row[5]), "estimasi_bulan": self.to_int(row[6]),
                        "gambar_path": row[7], "terkumpul": self.to_int(row[8]),
                        "notifikasi": row[9] or "-", "notif_aktif": row[10] or False,
                        "tanggal_dibuat": row[11]
                    })
        except:
            self.savings = []

    def save_user_data(self):
        if not self.user_manager.current_user or not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = load_workbook(EXCEL_FILE)
            if self.user_manager.current_user not in wb.sheetnames:
                self.create_user_sheet(self.user_manager.current_user)
                wb = load_workbook(EXCEL_FILE)
            ws = wb[self.user_manager.current_user]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            for item in self.savings:
                ws.append([item.get("nama"), item.get("target"), item.get("nominal"),
                           item.get("rencana"), item.get("estimasi_hari"),
                           item.get("estimasi_minggu"), item.get("estimasi_bulan"),
                           item.get("gambar_path"), item.get("terkumpul"),
                           item.get("notifikasi"), item.get("notif_aktif"),
                           item.get("tanggal_dibuat")])
            wb.save(EXCEL_FILE)
        except:
            pass

    def copy_image_for_user(self, source_path):
        if not os.path.exists(source_path) or not self.user_manager.current_user:
            return None
        try:
            images_folder = self.get_user_images_folder()
            if not images_folder:
                return None
            filename = os.path.basename(source_path)
            name, ext = os.path.splitext(filename)
            dest_path = os.path.join(images_folder, filename)
            counter = 1
            while os.path.exists(dest_path):
                dest_path = os.path.join(images_folder, f"{name}_{counter}{ext}")
                counter += 1
            shutil.copy2(source_path, dest_path)
            return dest_path
        except:
            return None

    def add_saving(self, data):
        self.savings.append(data)
        self.save_user_data()

    def update_saving(self, index, data):
        if 0 <= index < len(self.savings):
            self.savings[index] = data
            self.save_user_data()

    def delete_saving(self, index):
        if 0 <= index < len(self.savings):
            self.savings.pop(index)
            self.save_user_data()

    def get_filtered_savings(self, tab):
        if tab == "Berlangsung":
            return [d for d in self.savings if not self.is_completed(d)]
        else:
            return [d for d in self.savings if self.is_completed(d)]

    @staticmethod
    def to_int(val):
        try:
            return int(val)
        except:
            try:
                return int(float(val))
            except:
                return 0

    @staticmethod
    def is_completed(data):
        try:
            return int(data.get("terkumpul", 0)) >= int(data.get("target", 0))
        except:
            return False

class NotificationManager:
    def __init__(self, toaster):
        self.toaster = toaster
        self.last_notifications = {}
        self.running = mp.Value('b', True)
        self.process = mp.Process(target=self._notification_loop, args=(self.running,))
        self.process.start()

    def _notification_loop(self, running):
        while running.value:
            self._check_notifications()
            time.sleep(30)  # Cek setiap 30 detik

    def _check_notifications(self):
        current_user_file = os.path.join(DATA_FOLDER, "current_user.txt")
        if not os.path.exists(current_user_file):
            return
        try:
            with open(current_user_file, 'r') as f:
                current_user = f.read().strip()
        except:
            return
        if not current_user:
            return

        savings = []
        if os.path.exists(EXCEL_FILE):
            try:
                wb = load_workbook(EXCEL_FILE)
                if current_user in wb.sheetnames:
                    ws = wb[current_user]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            savings.append({
                                "nama": row[0], "target": self.to_int(row[1]), "nominal": self.to_int(row[2]),
                                "rencana": row[3] or "Harian", "estimasi_hari": self.to_int(row[4]),
                                "estimasi_minggu": self.to_int(row[5]), "estimasi_bulan": self.to_int(row[6]),
                                "gambar_path": row[7], "terkumpul": self.to_int(row[8]),
                                "notifikasi": row[9] or "-", "notif_aktif": row[10] or False,
                                "tanggal_dibuat": row[11]
                            })
            except:
                pass

        now_dt = datetime.now()
        now_h, now_m = now_dt.hour, now_dt.minute
        today = now_dt.date().isoformat()

        for data in savings:
            if not data.get("notif_aktif") or not data.get("notifikasi") or data.get("notifikasi") == "-":
                continue

            jam = data.get("notifikasi")
            if not isinstance(jam, str) or ":" not in jam:
                continue
            try:
                h, m = map(int, jam.split(":"))
            except:
                continue

            if h != now_h or m != now_m:
                continue

            key = f"{data.get('nama','')}|{data.get('tanggal_dibuat','')}"
            last = self.last_notifications.get(key)
            allowed = self._is_allowed_to_notify(data, last, today, now_dt)

            if not allowed:
                continue

            self.toaster.show_toast(
                "Pengingat Tabungan",
                f"Setor tabungan: {data.get('nama','')} — Rp {data.get('nominal',0):,}",
                duration=10,
                threaded=True
            )

            self.last_notifications[key] = today

    def _is_allowed_to_notify(self, data, last, today, now_dt):
        rencana = data.get('rencana', 'Harian')
        if rencana == 'Harian':
            return True
        elif rencana == 'Mingguan':
            if not last:
                return True
            try:
                last_dt = datetime.fromisoformat(last).date()
                return (now_dt.date() - last_dt).days >= 7
            except:
                return True
        elif rencana == 'Bulanan':
            if not last:
                return True
            try:
                last_dt = datetime.fromisoformat(last).date()
                return (now_dt.date() - last_dt).days >= 30
            except:
                return True
        else:
            return True

    @staticmethod
    def to_int(val):
        try:
            return int(val)
        except:
            try:
                return int(float(val))
            except:
                return 0

    def stop(self):
        self.running.value = False
        self.process.terminate()
        self.process.join()

class LoginFrame(ctk.CTkFrame):
    def __init__(self, master, user_manager, on_login_success, on_show_register):
        super().__init__(master, width=420, height=460, corner_radius=20)
        self.user_manager = user_manager
        self.on_login_success = on_login_success
        self.on_show_register = on_show_register
        self.setup_ui()

    def setup_ui(self):
        ctk.CTkLabel(self, text="LOGIN", font=("Poppins", 26, "bold")).pack(pady=10)

        self.username_entry = ctk.CTkEntry(self, width=300, height=45)
        self.username_entry.pack(pady=10)
        self.add_placeholder(self.username_entry, "Username")

        self.password_entry = ctk.CTkEntry(self, width=300, height=45)
        self.password_entry.pack(pady=10)
        self.add_placeholder(self.password_entry, "Password", is_password=True)

        ctk.CTkButton(self, text="Login", height=45, command=self.login_action).pack(pady=15)
        ctk.CTkButton(self, text="Belum punya akun?", fg_color="transparent", text_color="#3b82f6", hover=False, 
                      command=self.on_show_register).pack()

    def add_placeholder(self, entry, text, is_password=False):
        entry.delete(0, "end")
        entry.insert(0, text)
        entry.configure(text_color="#9ca3af")
        def on_in(_):
            if entry.get() == text:
                entry.delete(0, "end")
                entry.configure(text_color="#111827")
                if is_password: entry.configure(show="*")
        def on_out(_):
            if not entry.get().strip():
                entry.insert(0, text)
                entry.configure(text_color="#9ca3af")
                if is_password: entry.configure(show="")
        entry.bind("<FocusIn>", on_in)
        entry.bind("<FocusOut>", on_out)

    def login_action(self):
        uname = self.username_entry.get()
        pw = self.password_entry.get()
        if uname == "Username" or pw == "Password":
            messagebox.showerror("Error", "Isi semua field")
            return
        try:
            self.user_manager.login_user(uname, pw)
            self.on_login_success()
        except ValueError as e:
            messagebox.showerror("Error", str(e))

class RegisterFrame(ctk.CTkFrame):
    def __init__(self, master, user_manager, on_register_success, on_show_login):
        super().__init__(master, width=420, height=460, corner_radius=20)
        self.user_manager = user_manager
        self.on_register_success = on_register_success
        self.on_show_login = on_show_login
        self.setup_ui()

    def setup_ui(self):
        ctk.CTkLabel(self, text="REGISTER", font=("Poppins", 26, "bold")).pack(pady=10)

        self.reg_user_entry = ctk.CTkEntry(self, width=300, height=45)
        self.reg_user_entry.pack(pady=10)
        self.add_placeholder(self.reg_user_entry, "Username")

        self.reg_pass_entry = ctk.CTkEntry(self, width=300, height=45)
        self.reg_pass_entry.pack(pady=10)
        self.add_placeholder(self.reg_pass_entry, "Password", is_password=True)

        self.reg_confirm_entry = ctk.CTkEntry(self, width=300, height=45)
        self.reg_confirm_entry.pack(pady=10)
        self.add_placeholder(self.reg_confirm_entry, "Konfirmasi Password", is_password=True)

        ctk.CTkButton(self, text="Register", height=45, command=self.register_action).pack(pady=15)
        ctk.CTkButton(self, text="Udah punya akun?", fg_color="transparent", text_color="#3b82f6", hover=False, 
                      command=self.on_show_login).pack()

    def add_placeholder(self, entry, text, is_password=False):
        entry.delete(0, "end")
        entry.insert(0, text)
        entry.configure(text_color="#9ca3af")
        def on_in(_):
            if entry.get() == text:
                entry.delete(0, "end")
                entry.configure(text_color="#111827")
                if is_password: entry.configure(show="*")
        def on_out(_):
            if not entry.get().strip():
                entry.insert(0, text)
                entry.configure(text_color="#9ca3af")
                if is_password: entry.configure(show="")
        entry.bind("<FocusIn>", on_in)
        entry.bind("<FocusOut>", on_out)

    def register_action(self):
        uname = self.reg_user_entry.get()
        pw = self.reg_pass_entry.get()
        cp = self.reg_confirm_entry.get()
        if uname == "Username" or pw == "Password" or cp == "Konfirmasi Password":
            messagebox.showerror("Error", "Isi semua field")
            return  
        try:
            self.user_manager.register_user(uname, pw, cp)
            messagebox.showinfo("Success", "Registrasi berhasil")
            self.on_register_success()
        except ValueError as e:
            messagebox.showerror("Error", str(e))

class MainFrame(ctk.CTkFrame):
    def __init__(self, master, savings_manager, on_logout, on_add_saving, on_edit_saving):
        super().__init__(master, fg_color="#f9fafb")
        self.savings_manager = savings_manager
        self.on_logout = on_logout
        self.on_add_saving = on_add_saving
        self.on_edit_saving = on_edit_saving
        self.current_tab = "Berlangsung"
        self.setup_ui()

    def setup_ui(self):
        # HEADER
        header_frame = ctk.CTkFrame(self, fg_color="white", height=60, corner_radius=0)
        header_frame.pack(fill="x", pady=(0, 5))

        header_content = ctk.CTkFrame(header_frame, fg_color="white")
        header_content.pack(fill="both", expand=True, padx=20, pady=10)

        self.judul_label = ctk.CTkLabel(header_content, text="Savings Reminder", text_color="black", font=("Poppins", 22, "bold"), anchor="w")
        self.judul_label.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(header_content, text="Logout", width=80, height=40, fg_color="#ef4444", text_color="white", font=("Poppins", 12, "bold"), command=self.on_logout).pack(side="right", padx=5)

        # TAB
        tab_frame = ctk.CTkFrame(self, fg_color="white", corner_radius=0)
        tab_frame.pack(fill="x", pady=(0, 10))

        self.tab_berlangsung = ctk.CTkButton(tab_frame, text="Berlangsung", width=300, height=35, corner_radius=10, fg_color="#3b82f6", command=lambda: self.set_tab("Berlangsung"))
        self.tab_berlangsung.pack(side="left", padx=(50, 5), pady=5)

        self.tab_tercapai = ctk.CTkButton(tab_frame, text="Tercapai", width=300, height=35, corner_radius=10, fg_color="#e5e7eb", text_color="black", command=lambda: self.set_tab("Tercapai"))
        self.tab_tercapai.pack(side="left", padx=(5, 100), pady=5)

        # LIST CARD
        self.content_frame = ctk.CTkScrollableFrame(self, width=380, height=540, fg_color="transparent", scrollbar_fg_color="transparent", scrollbar_button_color="#d1d5db", scrollbar_button_hover_color="#9ca3af")
        self.content_frame.pack(pady=(5, 10), fill="both", expand=True)

        # FLOAT BUTTON
        float_frame = ctk.CTkFrame(self, fg_color="transparent")
        float_frame.place(relx=0.5, rely=0.92, anchor="center")

        ctk.CTkButton(float_frame, text="+ Tambah Celengan", width=220, height=45, corner_radius=25, font=("Poppins", 15, "bold"), command=self.on_add_saving).pack()

    def set_tab(self, tab):
        self.current_tab = tab
        self.update_tab_header()
        self.update_cards()

    def update_tab_header(self):
        if self.current_tab == "Berlangsung":
            self.tab_berlangsung.configure(fg_color="#3b82f6", text_color="white")
            self.tab_tercapai.configure(fg_color="#e5e7eb", text_color="black")
        else:
            self.tab_tercapai.configure(fg_color="#3b82f6", text_color="white")
            self.tab_berlangsung.configure(fg_color="#e5e7eb", text_color="black")

    def update_cards(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        list_filtered = self.savings_manager.get_filtered_savings(self.current_tab)
        if not list_filtered:
            ctk.CTkLabel(self.content_frame, text="Tidak ada data", font=("Poppins", 14), text_color="#6b7280").pack(pady=20)
            return
        for data in list_filtered:
            actual_index = self.savings_manager.savings.index(data)
            self.create_card(data, actual_index)

    def create_card(self, data, index):
        card = ctk.CTkFrame(self.content_frame, corner_radius=10, fg_color="white", border_width=1, border_color="#e5e7eb")
        card.pack(pady=8, padx=8, fill="x")

        def buka_detail_event():
            self.on_edit_saving(data, index)

        card.bind("<Button-1>", lambda e: buka_detail_event())

        # Judul
        ctk.CTkLabel(card, text=data["nama"], font=("Poppins", 16, "bold"), text_color="#111827").pack(pady=(6, 0))

        # Gambar
        img_frame = ctk.CTkFrame(card, width=250, height=100, fg_color="#f3f4f6", corner_radius=10)
        img_frame.pack(pady=6)

        if data.get("gambar_path"):
            try:
                img = Image.open(data["gambar_path"]).resize(IMAGE_SIZE_CARD)
                img_tk = ctk.CTkImage(light_image=img, size=IMAGE_SIZE_CARD)
                img_label = ctk.CTkLabel(img_frame, image=img_tk, text="")
                self.savings_manager.image_refs[id(img_label)] = img_tk
                img_label.pack()
            except:
                ctk.CTkLabel(img_frame, text="🖼️", font=("Arial", 38)).pack()
        else:
            ctk.CTkLabel(img_frame, text="🖼️", font=("Arial", 38)).pack()

        # Target
        ctk.CTkLabel(card, text=f"Target: Rp {data['target']:,}", font=("Poppins", 12)).pack(pady=(0, 6))

        # Info nominal & rencana
        info_frame = ctk.CTkFrame(card, fg_color="white", corner_radius=15)
        info_frame.pack(pady=(8,12), padx=15, fill="x")

        ctk.CTkLabel(info_frame, text=f"Rp {data.get('nominal',0):,} / {data.get('rencana','Harian')}", font=("Poppins", 13, "bold"), text_color="#111827").pack(pady=(6,2))

        # Estimasi
        if data.get("rencana") == "Harian":
            estimasi_text = f"{data.get('estimasi_hari',0)} Hari"
        elif data.get("rencana") == "Mingguan":
            estimasi_text = f"{data.get('estimasi_minggu',0)} Minggu"
        else:
            estimasi_text = f"{data.get('estimasi_bulan',0)} Bulan"

        ctk.CTkLabel(info_frame, text=f"Estimasi Tercapai: {estimasi_text}", font=("Poppins", 12), text_color="#2563eb").pack(pady=(0,8))

        # Progress bar
        terkumpul = data.get("terkumpul", 0)
        target = max(data.get("target", 1), 1)
        persen = (terkumpul / target) * 100

        progress_frame = ctk.CTkFrame(card, fg_color="white")
        progress_frame.pack(pady=(0,10), padx=15, fill="x")

        progress_bar = ctk.CTkProgressBar(progress_frame, height=12, corner_radius=8)
        progress_bar.pack(fill="x", padx=10, pady=(6,2))
        progress_bar.set(persen/100)

        ctk.CTkLabel(progress_frame, text=f"{persen:.1f}% Terkumpul", font=("Poppins", 11, "bold"), text_color="#16a34a").pack(pady=(0,6))

class InputFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, savings_manager, on_save, on_back):
        super().__init__(master, fg_color="#f9fafb", width=700, height=600)
        self.savings_manager = savings_manager
        self.on_save = on_save
        self.on_back = on_back
        self.edit_index = None
        self.selected_image = None
        self.setup_ui()

    def setup_ui(self):
        self.judul_input = ctk.CTkLabel(self, text="Tambah / Edit Celengan", font=("Poppins", 20, "bold"))
        self.judul_input.pack(pady=(20, 10))

        # FRAME GAMBAR
        self.gambar_frame = ctk.CTkFrame(self, width=300, height=150, fg_color="#e5e7eb", corner_radius=10)
        self.gambar_frame.pack(pady=10)

        self.gambar_label = ctk.CTkLabel(self.gambar_frame, text="🖼️ Tambah Gambar", font=("Poppins", 14))
        self.gambar_label.place(relx=0.5, rely=0.5, anchor="center")

        self.gambar_frame.bind("<Button-1>", lambda e: self.select_image())
        self.gambar_label.bind("<Button-1>", lambda e: self.select_image())

        # INPUT DASAR
        self.nama_entry = ctk.CTkEntry(self, placeholder_text="Nama Tabungan", width=300, height=40)
        self.nama_entry.pack(pady=10)

        self.target_entry = ctk.CTkEntry(self, placeholder_text="Target Tabungan (Rp)", width=300, height=40)
        self.target_entry.pack(pady=10)

        # RENCANA
        rencana_label = ctk.CTkLabel(self, text="Rencana Pengisian:", font=("Poppins", 14, "bold"))
        rencana_label.pack(pady=(15, 5))

        self.opsi_var = ctk.StringVar(value="Harian")

        rencana_frame = ctk.CTkFrame(self, fg_color="transparent")
        rencana_frame.pack()

        for opsi in ["Harian", "Mingguan", "Bulanan"]:
            ctk.CTkRadioButton(rencana_frame, text=opsi, variable=self.opsi_var, value=opsi).pack(side="left", padx=10)

        # NOMINAL
        self.nominal_entry = ctk.CTkEntry(self, placeholder_text="Nominal Pengisian", width=300, height=40)
        self.nominal_entry.pack(pady=(15, 5))

        # NOTIFIKASI
        notif_label = ctk.CTkLabel(self, text="Notifikasi", font=("Poppins", 14, "bold"))
        notif_label.pack(anchor="w", padx=15)

        notif_frame = ctk.CTkFrame(self, fg_color="white", corner_radius=10)
        notif_frame.pack(pady=10, padx=20, fill="x")

        self.time_var = ctk.StringVar(value="12:00")

        def ubah_waktu():
            jam_window = ctk.CTkToplevel(self)
            jam_window.title("Pilih Waktu")
            jam_window.geometry("200x180")

            ctk.CTkLabel(jam_window, text="Pilih Jam", font=("Poppins", 13, "bold")).pack(pady=5)

            jam_spin = ctk.CTkComboBox(jam_window, values=[f"{i:02d}" for i in range(24)], width=60)
            jam_spin.pack(pady=5)

            menit_spin = ctk.CTkComboBox(jam_window, values=[f"{i:02d}" for i in range(60)], width=60)
            menit_spin.pack(pady=5)

            def simpan_waktu():
                self.time_var.set(f"{jam_spin.get()}:{menit_spin.get()}")
                jam_window.destroy()

            ctk.CTkButton(jam_window, text="Pilih", command=simpan_waktu).pack(pady=10)

        time_label = ctk.CTkLabel(notif_frame, textvariable=self.time_var, font=("Poppins", 22, "bold"), text_color="#111827")
        time_label.pack(side="left", padx=10, pady=10)

        edit_waktu = ctk.CTkButton(notif_frame, text="✎", width=40, height=40, fg_color="#e5e7eb", text_color="black", command=ubah_waktu)
        edit_waktu.pack(side="left", padx=10)

        self.notif_switch_var = ctk.BooleanVar(value=False)
        notif_switch = ctk.CTkSwitch(notif_frame, variable=self.notif_switch_var, text="")
        notif_switch.pack(side="right", padx=10)

        # TOMBOL INPUT
        tombol_frame = ctk.CTkFrame(self, fg_color="transparent")
        tombol_frame.pack(pady=20)

        ctk.CTkButton(tombol_frame, text="Simpan", width=150, height=50, fg_color="#3b82f6", font=("Poppins", 14, "bold"), command=self.save_and_back).pack(side="left", padx=5)

        ctk.CTkButton(tombol_frame, text="Kembali", width=150, height=50, fg_color="#9ca3af", font=("Poppins", 14, "bold"), command=self.on_back).pack(side="left", padx=5)

    def select_image(self):
        file_path = filedialog.askopenfilename(title="Pilih Gambar", filetypes=[("File Gambar", "*.png;*.jpg;*.jpeg")])
        if file_path:
            self.selected_image = file_path
            try:
                img = Image.open(file_path).resize(IMAGE_SIZE_PREVIEW)
                img_tk = ctk.CTkImage(light_image=img, size=IMAGE_SIZE_PREVIEW)
                self.gambar_label.configure(image=img_tk, text="")
                self.savings_manager.image_refs[id(self.gambar_label)] = img_tk
            except Exception as e:
                print("Preview gambar gagal:", e)
                self.gambar_label.configure(text="🖼️ (preview gagal)")

    def load_data(self, data=None, index=None):
        self.edit_index = index
        self.selected_image = None

        self.nama_entry.delete(0, "end")
        self.target_entry.delete(0, "end")
        self.nominal_entry.delete(0, "end")
        self.opsi_var.set("Harian")
        self.gambar_label.configure(image=None, text="🖼️ Tambah Gambar")
        self.savings_manager.image_refs.pop(id(self.gambar_label), None)

        if data is not None:
            self.nama_entry.insert(0, data["nama"])
            self.target_entry.insert(0, str(data["target"]))
            self.opsi_var.set(data.get("rencana", "Harian"))
            self.nominal_entry.insert(0, str(data.get("nominal", "")))

            if data.get("gambar_path"):
                try:
                    img = Image.open(data["gambar_path"]).resize(IMAGE_SIZE_PREVIEW)
                    img_tk = ctk.CTkImage(light_image=img, size=IMAGE_SIZE_PREVIEW)
                    self.gambar_label.configure(image=img_tk, text="")
                    self.savings_manager.image_refs[id(self.gambar_label)] = img_tk
                    self.selected_image = data["gambar_path"]
                except:
                    pass

    def save_and_back(self):
        nama = self.nama_entry.get().strip()
        target = self.target_entry.get().strip()
        rencana = self.opsi_var.get()
        nominal = self.nominal_entry.get().strip()

        if not nama or not target or not nominal:
            messagebox.showerror("Error", "Semua field harus diisi!")
            return
        if not target.isdigit() or not nominal.isdigit():
            messagebox.showerror("Error", "Target & Nominal harus angka!")
            return

        nominal = int(nominal)
        target = int(target)

        notifikasi_waktu = self.time_var.get()
        notifikasi_status = self.notif_switch_var.get()

        estimasi_hari = target // nominal if nominal > 0 else 0
        estimasi_minggu = target // nominal
        estimasi_bulan = target // nominal

        if self.edit_index is not None and 0 <= self.edit_index < len(self.savings_manager.savings):
            existing = self.savings_manager.savings[self.edit_index]
            terkumpul_val = existing.get("terkumpul", 0)
            tanggal_dibuat = existing.get("tanggal_dibuat")
        else:
            terkumpul_val = 0
            tanggal_dibuat = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        # Copy gambar ke folder user jika ada gambar baru dipilih
        final_image_path = None
        if self.selected_image:
            final_image_path = self.savings_manager.copy_image_for_user(self.selected_image)

        # Jika edit, pertahankan gambar lama jika tidak ada gambar baru
        if self.edit_index is not None and 0 <= self.edit_index < len(self.savings_manager.savings) and not final_image_path:
            final_image_path = self.savings_manager.savings[self.edit_index].get("gambar_path")

        data = {
            "nama": nama,
            "target": target,
            "nominal": nominal,
            "estimasi_hari": estimasi_hari,
            "estimasi_minggu": estimasi_minggu,
            "estimasi_bulan": estimasi_bulan,
            "rencana": rencana,
            "gambar_path": final_image_path,
            "terkumpul": terkumpul_val,
            "notifikasi": notifikasi_waktu if notifikasi_status else "-",
            "notif_aktif": notifikasi_status,
            "tanggal_dibuat": tanggal_dibuat
        }

        if self.edit_index is not None and 0 <= self.edit_index < len(self.savings_manager.savings):
            self.savings_manager.update_saving(self.edit_index, data)
        else:
            self.savings_manager.add_saving(data)

        self.on_save()

class DetailFrame(ctk.CTkFrame):
    def __init__(self, master, savings_manager, on_back, on_edit):
        super().__init__(master, fg_color="#f9fafb")
        self.savings_manager = savings_manager
        self.on_back = on_back
        self.on_edit = on_edit
        self.detail_content = ctk.CTkScrollableFrame(self, fg_color="transparent", width=700, height=600)
        self.detail_content.pack(fill="both", expand=True)

    def load_detail(self, data, index):
        for widget in self.detail_content.winfo_children():
            widget.destroy()

        # Gambar
        if data.get("gambar_path"):
            try:
                img = Image.open(data["gambar_path"]).resize(IMAGE_SIZE_DETAIL)
                img_tk = ctk.CTkImage(light_image=img, size=IMAGE_SIZE_DETAIL)
                img_label = ctk.CTkLabel(self.detail_content, image=img_tk, text="")
                self.savings_manager.image_refs[id(img_label)] = img_tk
                img_label.pack(pady=10)
            except:
                ctk.CTkLabel(self.detail_content, text="🖼️", font=("Arial", 38)).pack(pady=10)
        else:
            ctk.CTkLabel(self.detail_content, text="🖼️", font=("Arial", 38)).pack(pady=10)

        # Bar 1: Nama + Target
        bar1 = ctk.CTkFrame(self.detail_content, fg_color="white", corner_radius=10)
        bar1.pack(pady=5, padx=15, fill="x")

        nama_label = ctk.CTkLabel(bar1, text=f"Nama: {data['nama']}", font=("Poppins", 16, "bold"), text_color="#111827")
        nama_label.pack(side="left", padx=10, pady=8)

        target_label = ctk.CTkLabel(bar1, text=f"Harga: Rp {data['target']:,}", font=("Poppins", 14), text_color="#2563eb")
        target_label.pack(side="right", padx=10)

        # Bar 2: Tanggal & Estimasi
        bar2 = ctk.CTkFrame(self.detail_content, fg_color="white", corner_radius=10)
        bar2.pack(pady=5, padx=15, fill="x")

        tanggal = datetime.now().strftime("%d %B %Y")

        if data.get("rencana") == "Harian":
            estimasi_str = f"{data.get('estimasi_hari', 0)} Hari"
        elif data.get("rencana") == "Mingguan":
            estimasi_str = f"{data.get('estimasi_minggu', 0)} Minggu"
        else:
            estimasi_str = f"{data.get('estimasi_bulan', 0)} Bulan"

        tanggal_label = ctk.CTkLabel(bar2, text=f"📅 Dibuat: {tanggal}", font=("Poppins", 13))
        tanggal_label.pack(side="left", padx=10, pady=8)

        estimasi_label = ctk.CTkLabel(bar2, text=f"⏳ Estimasi: {estimasi_str}", font=("Poppins", 13))
        estimasi_label.pack(side="right", padx=10)

        # Bar 3: Waktu submit
        bar3 = ctk.CTkFrame(self.detail_content, fg_color="white", corner_radius=10)
        bar3.pack(pady=5, padx=15, fill="x")

        waktu = (datetime.now() + timedelta(days=1)).strftime("%H:%M")
        ctk.CTkLabel(bar3, text=f"Waktu: {waktu}", font=("Poppins", 13)).pack(padx=10, pady=8, anchor="w")

        notif = data.get("notifikasi", "-")
        notif_label = ctk.CTkLabel(bar3, text=f"🔔 Notifikasi: {notif}", font=("Poppins", 13), text_color="#ea580c")
        notif_label.pack(side="left", padx=10, pady=(5), anchor="w")

        # Bar 4: Terkumpul & Kekurangan
        bar4 = ctk.CTkFrame(self.detail_content, fg_color="white", corner_radius=10)
        bar4.pack(pady=10, padx=15, fill="x")

        terkumpul = data.get("terkumpul", 0)
        kekurangan = data["target"] - terkumpul
        if kekurangan < 0:
            kekurangan = 0

        left = ctk.CTkFrame(bar4, fg_color="white")
        left.pack(side="left", expand=True, fill="both", padx=(0, 2), pady=5)

        right = ctk.CTkFrame(bar4, fg_color="white")
        right.pack(side="left", expand=True, fill="both", padx=(2, 0), pady=5)

        ctk.CTkLabel(left, text="Terkumpul", font=("Poppins", 13, "bold"), text_color="#16a34a").pack()
        terkumpul_label = ctk.CTkLabel(left, text=f"Rp {terkumpul:,}", font=("Poppins", 14))
        terkumpul_label.pack()

        ctk.CTkLabel(right, text="Kekurangan", font=("Poppins", 13, "bold"), text_color="#dc2626").pack()
        kekurangan_label = ctk.CTkLabel(right, text=f"Rp {kekurangan:,}", font=("Poppins", 14))
        kekurangan_label.pack()

        line = ctk.CTkFrame(bar4, width=1, fg_color="#d1d5db")
        line.place(relx=0.5, rely=0.5, anchor="center", relheight=0.7)

        # Area input: setor nominal
        isi_frame = ctk.CTkFrame(self.detail_content, fg_color="transparent")
        isi_frame.pack(pady=(8,12), padx=15, fill="x")

        ctk.CTkLabel(isi_frame, text="Masukkan Nominal Setor:", font=("Poppins", 13)).pack(anchor="w", padx=6)
        setor_entry = ctk.CTkEntry(isi_frame, placeholder_text="Nominal (Rp)", width=200, height=35)
        setor_entry.pack(side="left", padx=(6,8), pady=8)

        def tambah_setor():
            nonlocal terkumpul, kekurangan
            val = setor_entry.get().strip()
            if not val:
                messagebox.showerror("Error", "Masukkan nominal yang ingin disetor.")
                return
            if val.isalpha():
                messagebox.showerror("Error", "Nominal harus berupa angka.")
                return
            jumlah = int(val)
            if jumlah < 0:
                messagebox.showerror("Error", "Nominal harus lebih besar dari 0.")
                return
            if jumlah == 0:
                messagebox.showerror("Error", "Nominal harus lebih besar dari 0.")
                return

            self.savings_manager.savings[index]["terkumpul"] = self.savings_manager.savings[index].get("terkumpul", 0) + jumlah
            self.savings_manager.save_user_data()

            terkumpul = self.savings_manager.savings[index]["terkumpul"]
            kekurangan = self.savings_manager.savings[index]["target"] - terkumpul
            if kekurangan < 0:
                kekurangan = 0

            terkumpul_label.configure(text=f"Rp {terkumpul:,}")
            kekurangan_label.configure(text=f"Rp {kekurangan:,}")

            setor_entry.delete(0, "end")

        ctk.CTkButton(isi_frame, text="Tambah", width=90, height=35, fg_color="#16a34a", command=tambah_setor).pack(side="left", padx=(0,6))

        # Tombol
        tombol_frame = ctk.CTkFrame(self.detail_content, fg_color="transparent")
        tombol_frame.pack(pady=20)

        ctk.CTkButton(tombol_frame, text="Edit Data", width=120, height=40, fg_color="#3b82f6", font=("Poppins", 14, "bold"), command=lambda: self.on_edit(data, index)).pack(side="left", padx=5)

        def hapus_data():
            ask = messagebox.askyesno("Hapus Data", f"Yakin ingin menghapus tabungan '{data['nama']}'?")
            if ask:
                self.savings_manager.delete_saving(index)
                self.on_back()

        ctk.CTkButton(tombol_frame, text="Hapus", width=120, height=40, fg_color="#ef4444", text_color="white", font=("Poppins", 14, "bold"), command=hapus_data).pack(side="left", padx=5)

        ctk.CTkButton(tombol_frame, text="Kembali", width=120, height=40, fg_color="#9ca3af", font=("Poppins", 14, "bold"), command=self.on_back).pack(side="left", padx=5)

class SavingsApp:
    def __init__(self):
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.app = ctk.CTk()
        self.app.title("Savings Reminder")
        self.app.geometry("700x600")
        self.app.resizable(False, False)

        self.toaster = ToastNotifier()

        self.user_manager = UserManager()
        self.savings_manager = SavingsManager(self.user_manager)
        self.notification_manager = NotificationManager(self.toaster)

        self.login_frame = LoginFrame(self.app, self.user_manager, self.on_login_success, self.show_register)
        self.register_frame = RegisterFrame(self.app, self.user_manager, self.show_login, self.show_login)
        self.main_frame = MainFrame(self.app, self.savings_manager, self.logout, self.show_input, self.show_detail)
        self.input_frame = InputFrame(self.app, self.savings_manager, self.back_to_main, self.back_to_main)
        self.detail_frame = DetailFrame(self.app, self.savings_manager, self.back_to_main, self.show_input)

    def show_login(self):
        self.register_frame.pack_forget()
        self.login_frame.pack(pady=40)

    def show_register(self):
        self.login_frame.pack_forget()
        self.register_frame.pack(pady=40)

    def on_login_success(self):
        self.savings_manager.load_user_data()
        self.login_frame.pack_forget()
        self.register_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)
        self.main_frame.judul_label.configure(text=f"Savings Reminder - {self.user_manager.current_user}")
        self.main_frame.update_tab_header()
        self.main_frame.update_cards()

    def logout(self):
        self.savings_manager.save_user_data()
        self.user_manager.logout_user()
        self.savings_manager.savings.clear()
        self.main_frame.pack_forget()
        self.input_frame.pack_forget()
        self.detail_frame.pack_forget()
        self.notification_manager.stop()
        self.show_login()

    def show_input(self, data=None, index=None):
        self.main_frame.pack_forget()
        self.detail_frame.pack_forget()
        self.input_frame.pack(fill="both", expand=True)
        self.input_frame.load_data(data, index)

    def show_detail(self, data, index):
        self.main_frame.pack_forget()
        self.input_frame.pack_forget()
        self.detail_frame.pack(fill="both", expand=True)
        self.detail_frame.load_detail(data, index)

    def back_to_main(self):
        self.input_frame.pack_forget()
        self.detail_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)
        self.main_frame.update_cards()

    def run(self):
        self.show_login()
        self.app.mainloop()

if __name__ == "__main__":
    app = SavingsApp()
    app.run()

